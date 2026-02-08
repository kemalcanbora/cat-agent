"""Spreadsheet parsers: Excel (.xlsx/.xls), CSV and TSV."""

import re
from typing import List

from cat_agent.log import logger


# ---------------------------------------------------------------------------
# DataFrame → Markdown conversion
# ---------------------------------------------------------------------------


def df_to_md(df) -> str:
    """Convert a Polars DataFrame to a markdown table string."""
    import polars as pl

    def _replace_long_dashes(text):
        if text.replace('-', '').replace(':', '').strip():
            return text
        return re.sub(r'-{6,}', '-----', text)

    # Drop all-null columns and all-null rows
    non_null_cols = [col for col in df.columns if df[col].null_count() < len(df)]
    df = df.select(non_null_cols) if non_null_cols else df
    df = df.filter(~pl.all_horizontal(pl.all().is_null()))

    # Cast everything to string and fill nulls
    df = df.cast({col: pl.Utf8 for col in df.columns})
    df = df.fill_null('')

    headers = df.columns
    header_row = '| ' + ' | '.join(headers) + ' |'
    separator_row = '|' + '|'.join([' ----- ' for _ in headers]) + '|'

    data_rows = []
    for row in df.iter_rows():
        data_rows.append('| ' + ' | '.join(str(val) for val in row) + ' |')

    md_table = '\n'.join([header_row, separator_row] + data_rows)

    # Clean up long dash sequences
    md_table = '\n'.join([
        '|'.join(
            _replace_long_dashes(' ' + cell.strip() + ' ' if cell else '')
            for cell in row.split('|'))
        for row in md_table.split('\n')
    ])
    return md_table


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def parse_excel(file_path: str, extract_image: bool = False) -> List[dict]:
    if extract_image:
        raise ValueError('Currently, extracting images is not supported!')

    import polars as pl
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    md_tables = []
    for sheet_name in sheet_names:
        df = pl.read_excel(file_path, sheet_name=sheet_name)
        md_table = df_to_md(df)
        md_tables.append(f'### Sheet: {sheet_name}\n{md_table}')

    return [{'page_num': i + 1, 'content': [{'table': md_tables[i]}]} for i in range(len(md_tables))]


# ---------------------------------------------------------------------------
# CSV / TSV
# ---------------------------------------------------------------------------


def parse_csv(file_path: str, extract_image: bool = False) -> List[dict]:
    if extract_image:
        raise ValueError('Currently, extracting images is not supported!')

    import polars as pl

    try:
        df = pl.read_csv(file_path, ignore_errors=True, truncate_ragged_lines=True)
    except Exception as ex:
        logger.warning(ex)
        return parse_excel(file_path, extract_image)

    md_table = df_to_md(df)
    return [{'page_num': 1, 'content': [{'table': md_table}]}]


def parse_tsv(file_path: str, extract_image: bool = False) -> List[dict]:
    if extract_image:
        raise ValueError('Currently, extracting images is not supported!')

    import polars as pl

    try:
        df = pl.read_csv(file_path, separator='\t', ignore_errors=True, truncate_ragged_lines=True)
    except Exception as ex:
        logger.warning(ex)
        return parse_excel(file_path, extract_image)

    md_table = df_to_md(df)
    return [{'page_num': 1, 'content': [{'table': md_table}]}]
