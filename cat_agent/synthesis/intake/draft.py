"""Load a business-user Markdown (or Excel) draft into a :class:`Draft`."""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple, Union

from cat_agent.log import logger
from cat_agent.synthesis.intake.lang import detect_lang
from cat_agent.synthesis.intake.numbers import (
    CellParse,
    parse_cell,
    parse_column_values,
)
from cat_agent.synthesis.spec import Example

# GitHub-flavoured markdown table rows.
_TABLE_ROW = re.compile(r'^\s*\|(.+)\|\s*$')
_SEP_ROW = re.compile(r'^\s*\|[\s:\-|]+\|\s*$')

_LOCALE_SECTION = re.compile(
    r'(?im)^#{1,3}\s*(locale|sprache|langue|idioma|dil|taal)\s*$'
)


@dataclass
class OpenQuestion:
    """Something Task 2 could not resolve without guessing."""

    kind: str  # 'number' | 'date' | 'table' | 'other'
    message: str
    raw: str = ''
    row: Optional[int] = None
    column: Optional[str] = None


@dataclass
class Draft:
    raw_markdown: str
    source_path: Optional[str]
    examples: List[Example]
    example_columns: List[str]
    locale: Optional[str] = None
    open_questions: List[OpenQuestion] = field(default_factory=list)
    table_ambiguity: Optional[str] = None
    detected_lang: str = 'en'

    @classmethod
    def from_markdown(
        cls,
        text: str,
        *,
        source_path: Optional[str] = None,
        locale: Optional[str] = None,
    ) -> 'Draft':
        locale = locale or _extract_locale(text)
        tables = extract_markdown_tables(text)
        if not tables:
            raise ValueError(
                'No examples table found in the draft. Add a Markdown table with '
                'at least 3 data rows; the last column is the expected result. '
                'See: cat-agent synth init <name>'
            )
        if len(tables) > 1:
            # Largest by data-row count; note ambiguity for the interview.
            tables_sorted = sorted(tables, key=lambda t: len(t[1]), reverse=True)
            headers, rows = tables_sorted[0]
            note = (
                f'The draft contains {len(tables)} tables; using the largest '
                f'({len(rows)} data rows, columns={headers!r}).'
            )
        else:
            headers, rows = tables[0]
            note = None

        if len(headers) < 2:
            raise ValueError(
                'Examples table needs at least one input column and a result column.'
            )
        if len(rows) < 3:
            raise ValueError(
                f'Examples table has {len(rows)} data row(s); provide at least 3.'
            )

        examples, open_qs = _rows_to_examples(headers, rows, locale=locale)
        draft = cls(
            raw_markdown=text,
            source_path=source_path,
            examples=examples,
            example_columns=list(headers),
            locale=locale,
            open_questions=open_qs,
            table_ambiguity=note,
            detected_lang=detect_lang(text),
        )
        if note:
            draft.open_questions.append(
                OpenQuestion(kind='table', message=note),
            )
        return draft

    @classmethod
    def from_path(
        cls,
        path: Union[str, Path],
        *,
        locale: Optional[str] = None,
    ) -> 'Draft':
        file_path = Path(path)
        text = file_path.read_text(encoding='utf-8')
        return cls.from_markdown(text, source_path=str(file_path), locale=locale)

    @classmethod
    def from_excel(
        cls,
        path: Union[str, Path],
        description_md: Optional[str] = None,
        *,
        locale: Optional[str] = None,
    ) -> 'Draft':
        """Convert the first sheet to a markdown table and reuse :meth:`from_markdown`.

        Prefers native typed cell values from openpyxl when available so locale
        re-parsing is unnecessary for numeric cells.
        """
        file_path = Path(path)
        md_table, native_grid = _excel_to_markdown_and_natives(file_path)
        prose = (description_md or '').rstrip()
        if prose:
            text = f'{prose}\n\n## Examples\n{md_table}\n'
        else:
            text = f'# tool\n\n## Examples\n{md_table}\n'

        if native_grid is not None:
            headers = [str(h) for h in native_grid[0]]
            data_rows = native_grid[1:]
            locale = locale or _extract_locale(text)
            examples, open_qs = _rows_to_examples(
                headers, data_rows, locale=locale, cells_are_native=True)
            return cls(
                raw_markdown=text,
                source_path=str(file_path),
                examples=examples,
                example_columns=headers,
                locale=locale,
                open_questions=open_qs,
                detected_lang=detect_lang(text),
            )
        return cls.from_markdown(text, source_path=str(file_path), locale=locale)


def extract_markdown_tables(text: str) -> List[Tuple[List[str], List[List[str]]]]:
    """Return ``[(headers, rows), ...]`` for every GFM table in *text*."""
    lines = text.splitlines()
    tables: List[Tuple[List[str], List[List[str]]]] = []
    i = 0
    while i < len(lines):
        if not _TABLE_ROW.match(lines[i]):
            i += 1
            continue
        header_cells = _split_row(lines[i])
        if i + 1 >= len(lines) or not _SEP_ROW.match(lines[i + 1]):
            i += 1
            continue
        i += 2
        rows: List[List[str]] = []
        while i < len(lines) and _TABLE_ROW.match(lines[i]):
            if _SEP_ROW.match(lines[i]):
                break
            rows.append(_split_row(lines[i]))
            i += 1
        if header_cells:
            # Normalise ragged rows
            width = len(header_cells)
            norm_rows = []
            for row in rows:
                padded = list(row) + [''] * max(0, width - len(row))
                norm_rows.append(padded[:width])
            tables.append((header_cells, norm_rows))
        continue
    return tables


def _split_row(line: str) -> List[str]:
    inner = line.strip()
    if inner.startswith('|'):
        inner = inner[1:]
    if inner.endswith('|'):
        inner = inner[:-1]
    return [c.strip() for c in inner.split('|')]


def _extract_locale(text: str) -> Optional[str]:
    lines = text.splitlines()
    for index, line in enumerate(lines):
        if _LOCALE_SECTION.match(line.strip()):
            # Next non-empty non-heading line
            for follow in lines[index + 1:]:
                stripped = follow.strip()
                if not stripped:
                    continue
                if stripped.startswith('#'):
                    break
                if stripped.startswith('<'):
                    return None
                # Take first token that looks like a locale tag
                token = stripped.split()[0].strip('.,;')
                if re.fullmatch(r'[A-Za-z]{2}([-_][A-Za-z]{2})?', token):
                    return token.replace('_', '-')
                return stripped[:32]
    return None


def _rows_to_examples(
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    locale: Optional[str],
    cells_are_native: bool = False,
) -> Tuple[List[Example], List[OpenQuestion]]:
    input_cols = list(headers[:-1])
    result_col = headers[-1]
    open_qs: List[OpenQuestion] = []
    examples: List[Example] = []

    # Parse column-wise so we can detect per-column ambiguity patterns
    columns: Dict[str, List[Any]] = {h: [] for h in headers}
    for row in rows:
        for h, cell in zip(headers, row):
            columns[h].append(cell)

    parsed_cols: Dict[str, List[Any]] = {}
    for h, cells in columns.items():
        values, report = parse_column_values(cells, locale=locale)
        # When cells are already native (Excel), re-parse with prefer_native
        if cells_are_native:
            values = []
            report.ambiguous.clear()
            for idx, cell in enumerate(cells):
                cp = parse_cell(cell, locale=locale, prefer_native=True)
                if cp.ambiguous:
                    report.ambiguous.append((idx, cp.raw, cp.question or ''))
                    values.append(cp.raw)
                else:
                    values.append(cp.value)
        parsed_cols[h] = values
        for row_i, raw, question in report.ambiguous:
            open_qs.append(
                OpenQuestion(
                    kind='number' if 'day-first' not in (question or '') else 'date',
                    message=question,
                    raw=raw,
                    row=row_i,
                    column=h,
                )
            )

    for row_i in range(len(rows)):
        # Skip completely empty rows
        if all(
            (parsed_cols[h][row_i] is None or parsed_cols[h][row_i] == '')
            for h in headers
        ):
            continue
        inputs = {col: parsed_cols[col][row_i] for col in input_cols}
        expected = parsed_cols[result_col][row_i]
        # Try JSON object/array in result cell when it is still a string
        if isinstance(expected, str):
            expected = _maybe_json(expected)
        examples.append(Example(inputs=inputs, expected=expected))

    return examples, open_qs


def _maybe_json(text: str) -> Any:
    s = text.strip()
    if not s:
        return None
    if s[0] in '{[':
        try:
            return json.loads(s)
        except json.JSONDecodeError:
            # Allow single-quoted pseudo-JSON from business users? No — keep literal.
            return text
    return text


def _excel_to_markdown_and_natives(
    path: Path,
) -> Tuple[str, Optional[List[List[Any]]]]:
    """Return (markdown_table, native_grid_or_None)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        from cat_agent.tools.parsers.excel_parser import parse_excel
        pages = parse_excel(str(path))
        content = pages[0]['content'][0]['table']
        # Strip ### Sheet header if present
        lines = content.splitlines()
        if lines and lines[0].startswith('###'):
            content = '\n'.join(lines[1:]).lstrip()
        return content, None

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        grid: List[List[Any]] = []
        for row in ws.iter_rows(values_only=True):
            if row is None or all(c is None for c in row):
                continue
            grid.append(list(row))
    finally:
        wb.close()

    if not grid:
        raise ValueError(f'Excel file {path} has no data rows.')

    # Build markdown from stringified values for raw_markdown provenance
    from cat_agent.tools.parsers.excel_parser import df_to_md
    try:
        import polars as pl
        headers = [str(c) if c is not None else f'col{i}' for i, c in enumerate(grid[0])]
        data = [
            {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
            for row in grid[1:]
        ]
        df = pl.DataFrame(data)
        md = df_to_md(df)
    except Exception as exc:
        logger.warning('Falling back to manual markdown for Excel: {}', exc)
        headers = [str(c) if c is not None else f'col{i}' for i, c in enumerate(grid[0])]
        lines = [
            '| ' + ' | '.join(headers) + ' |',
            '|' + '|'.join(['---'] * len(headers)) + '|',
        ]
        for row in grid[1:]:
            cells = ['' if c is None else str(c) for c in row]
            cells += [''] * max(0, len(headers) - len(cells))
            lines.append('| ' + ' | '.join(cells[:len(headers)]) + ' |')
        md = '\n'.join(lines)
        return md, grid

    # Replace header row of grid with string headers for consistency
    native = [headers] + grid[1:]
    return md, native
